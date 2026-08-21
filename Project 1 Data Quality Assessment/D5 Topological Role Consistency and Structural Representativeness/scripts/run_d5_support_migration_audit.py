from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

import matplotlib as mpl
import matplotlib.colors as mcolors
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


D5_ROOT = Path(__file__).resolve().parents[1]
SRC = D5_ROOT / "src"
if str(SRC) not in sys.path:
    sys.path.insert(0, str(SRC))

from d5_local.figures.figure_style import (  # noqa: E402
    PALETTE,
    PROFILE,
    configure_style,
    panel_label,
    style_axes,
)
from d5_local.publication.support_migration import (  # noqa: E402
    D5SupportMigrationAudit,
    write_manifest,
)


SHEET_NAMES = {
    "01_monthly_regime_occupancy": "01_monthly_occupancy",
    "01b_pre_post_regime_shift": "01b_pre_post_shift",
    "02_template_occupancy_56": "02_template_occupancy_56",
    "03_L1_to_L2_blockers": "03_L1_L2_blockers",
    "04_L2_to_L3_blockers": "04_L2_L3_blockers",
    "05_coverage_loss_attribution": "05_loss_attribution",
    "06_counterfactual_coverage": "06_counterfactual",
    "07_reference_horizon_sensitivity": "07_reference_horizon",
    "08_monthly_report_eligibility": "08_monthly_eligibility",
}

FIGURE_WIDTH_MM = 183.0
RASTER_DPI = 600
mpl.rcParams.update({
    "svg.fonttype": "none",
    "pdf.fonttype": 42,
})


def _save_audit_figure(fig: plt.Figure, output_root: Path, stem: str) -> list[Path]:
    output_root.mkdir(parents=True, exist_ok=True)
    svg = output_root / f"{stem}.svg"
    pdf = output_root / f"{stem}.pdf"
    png = output_root / f"{stem}.png"
    tiff = output_root / f"{stem}.tiff"
    fig.savefig(svg, facecolor="white")
    svg.write_text(
        "\n".join(line.rstrip() for line in svg.read_text(encoding="utf-8").splitlines()) + "\n",
        encoding="utf-8",
    )
    fig.savefig(pdf, facecolor="white")
    fig.savefig(png, dpi=RASTER_DPI, facecolor="white")
    fig.savefig(
        tiff,
        dpi=RASTER_DPI,
        facecolor="white",
        pil_kwargs={"compression": "tiff_lzw"},
    )
    plt.close(fig)
    return [png, pdf, svg, tiff]


def _write_workbook(path: Path, tables: dict[str, pd.DataFrame]) -> None:
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, frame in tables.items():
            frame.to_excel(writer, sheet_name=SHEET_NAMES[name], index=False)
        for sheet in writer.book.worksheets:
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            for cell in sheet[1]:
                cell.font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="3D5A80")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            sheet.row_dimensions[1].height = 30
            for column_index, column in enumerate(sheet.columns, start=1):
                values = [str(cell.value) if cell.value is not None else "" for cell in column]
                width = min(max(max(map(len, values), default=0) + 2, 11), 36)
                sheet.column_dimensions[get_column_letter(column_index)].width = width
                for cell in column[1:]:
                    cell.font = Font(name="Arial", size=9)
                    cell.alignment = Alignment(vertical="center", wrap_text=False)


def _write_table_files(output_root: Path, tables: dict[str, pd.DataFrame]) -> list[Path]:
    artifacts: list[Path] = []
    for name, frame in tables.items():
        parquet = output_root / f"{name}.parquet"
        csv = output_root / f"{name}.csv"
        frame.to_parquet(parquet, index=False)
        frame.to_csv(csv, index=False, encoding="utf-8-sig")
        artifacts.extend([parquet, csv])
    return artifacts


def _month_label(value: str) -> str:
    timestamp = pd.Period(value, freq="M").to_timestamp()
    return timestamp.strftime("%b\n%Y")


def _figure_occupancy(
    monthly: pd.DataFrame,
    figure_root: Path,
    reference_end: pd.Timestamp,
) -> list[Path]:
    configure_style()
    pivot = monthly.pivot(index="month", columns="regime_label", values="occupancy_rate")
    pivot = pivot.reindex(columns=["R0", "R1", "R2", "R3"], fill_value=0.0).fillna(0.0)
    ood = monthly.drop_duplicates("month").set_index("month")["ood_rate"].reindex(pivot.index)
    fig, ax = plt.subplots(figsize=(FIGURE_WIDTH_MM / 25.4, 3.35), constrained_layout=True)
    colors = [PALETTE["navy"], PALETTE["teal"], PALETTE["gold"], PALETTE["red"]]
    bottom = np.zeros(len(pivot))
    x = np.arange(len(pivot))
    for regime, color in zip(pivot.columns, colors):
        values = pivot[regime].to_numpy()
        ax.bar(x, values * 100.0, bottom=bottom * 100.0, color=color, width=0.72, label=regime)
        bottom += values
    ax.plot(x, ood.to_numpy() * 100.0, color=PALETTE["dark"], marker="o", ms=3.2,
            linestyle="--", linewidth=1.0, label="OOD")
    reference_month = reference_end.to_period("M").strftime("%Y-%m")
    month_index = list(pivot.index).index(reference_month)
    elapsed_month = (
        (reference_end.day - 1)
        + reference_end.hour / 24.0
        + reference_end.minute / (24.0 * 60.0)
    ) / reference_end.days_in_month
    reference_x = month_index - 0.5 + elapsed_month
    ax.axvline(reference_x, color=PALETTE["gray"], linewidth=0.8, linestyle=":")
    ax.text(
        reference_x - 0.08,
        104,
        f"Reference freeze: {reference_end:%d %b %Y}",
        ha="right",
        va="bottom",
        color=PALETTE["gray"],
    )
    ax.set_ylabel("Plant-global occupancy (%)")
    ax.set_xlabel("Month")
    ax.set_xticks(x, [_month_label(value) for value in pivot.index])
    ax.set_ylim(0, 112)
    handles, labels = ax.get_legend_handles_labels()
    order = [labels.index(label) for label in ["R0", "R1", "R2", "R3", "OOD"]]
    ax.legend(
        [handles[index] for index in order],
        [labels[index] for index in order],
        ncol=5,
        loc="upper left",
        bbox_to_anchor=(0.0, 1.02),
    )
    style_axes(ax)
    return _save_audit_figure(fig, figure_root, "FigD5_S1_monthly_regime_occupancy_migration")


def _sensor_sort_key(sensor: str) -> tuple[int, int, int]:
    analyte, line, position = sensor.split("_")
    return (0 if analyte == "DO" else 1, int(line), int(position))


def _figure_maturity(templates: pd.DataFrame, figure_root: Path) -> list[Path]:
    configure_style()
    sensors = sorted(templates["sensor_id"].unique(), key=_sensor_sort_key)
    levels = templates.pivot(index="sensor_id", columns="regime_label", values="support_level")
    levels = levels.reindex(index=sensors, columns=["R0", "R1", "R2", "R3"])
    numeric = levels.apply(
        lambda column: column.map({"L0": 0, "L1": 1, "L2": 2, "L3": 3})
    ).astype(float)
    shares = templates.pivot(index="sensor_id", columns="regime_label", values="post_ref_share")
    shares = shares.reindex(index=sensors, columns=["R0", "R1", "R2", "R3"]).fillna(0.0)
    cmap = mcolors.ListedColormap(["#E5E7EB", "#E9C46A", "#8ECAE6", "#2A9D8F"])
    norm = mcolors.BoundaryNorm([-0.5, 0.5, 1.5, 2.5, 3.5], cmap.N)
    fig, ax = plt.subplots(figsize=(FIGURE_WIDTH_MM / 25.4, 4.75), constrained_layout=True)
    image = ax.imshow(numeric.to_numpy(), cmap=cmap, norm=norm, aspect="auto")
    for row in range(len(sensors)):
        for column in range(4):
            label = levels.iloc[row, column]
            ax.text(column, row, label, ha="center", va="center", fontsize=6.4,
                    fontweight="bold" if label == "L1" else "normal")
            share = float(shares.iloc[row, column])
            if share > 0:
                ax.scatter(column, row, s=30 + 950 * share, facecolors="none",
                           edgecolors=PALETTE["dark"], linewidths=0.8)
    ax.set_xticks(range(4), ["R0", "R1", "R2", "R3"])
    ax.set_yticks(range(len(sensors)), sensors)
    ax.set_xlabel("Frozen process-regime template")
    ax.set_ylabel("Sensor")
    ax.tick_params(top=False, right=False)
    for spine in ax.spines.values():
        spine.set_linewidth(PROFILE.axis_line_pt)
    colorbar = fig.colorbar(image, ax=ax, fraction=0.026, pad=0.02, ticks=[0, 1, 2, 3])
    colorbar.ax.set_yticklabels(["L0", "L1", "L2", "L3"])
    colorbar.set_label("Evidence maturity")
    ax.text(1.0, -0.13, "Circle area: post-reference occupancy", transform=ax.transAxes,
            ha="right", va="top", color=PALETTE["gray"])
    return _save_audit_figure(fig, figure_root, "FigD5_S2_sensor_regime_maturity_map")


def _figure_blockers(blockers: pd.DataFrame, figure_root: Path) -> list[Path]:
    configure_style()
    blockers = blockers.sort_values(["analyte", "sensor_id"]).reset_index(drop=True)
    columns = ["family_days", "family_months", "node_days", "node_months", "node_coverage"]
    labels = ["Family\ndays", "Family\nmonths", "Node\ndays", "Node\nmonths", "Node\ncoverage"]
    matrix = blockers[columns].astype(int).to_numpy()
    fig, (ax, bar_ax) = plt.subplots(
        1, 2, figsize=(FIGURE_WIDTH_MM / 25.4, 4.85),
        gridspec_kw={"width_ratios": [3.1, 1.25]}, constrained_layout=True
    )
    cmap = mcolors.ListedColormap(["#F5F6F7", PALETTE["red"]])
    ax.imshow(matrix, cmap=cmap, vmin=0, vmax=1, aspect="auto")
    for row in range(len(blockers)):
        for column in range(len(columns)):
            ax.text(column, row, "Blocked" if matrix[row, column] else "Pass",
                    ha="center", va="center", fontsize=5.8,
                    color="white" if matrix[row, column] else PALETTE["gray"])
    ax.set_xticks(range(len(columns)), labels)
    ax.set_yticks(range(len(blockers)), blockers["sensor_id"])
    ax.set_ylabel("L1 sensor-regime template")
    ax.tick_params(axis="x", pad=4)
    for spine in ax.spines.values():
        spine.set_linewidth(PROFILE.axis_line_pt)
    panel_label(ax, "a")

    analyte_loss = blockers.groupby("analyte", observed=True)[
        "support_attributable_loss_hours"
    ].sum()
    analyte_loss = analyte_loss.reindex(["DO", "ORP"])
    y = np.arange(len(analyte_loss))
    hours = analyte_loss.to_numpy()
    bar_ax.barh(y, hours, color=[PALETTE["blue"], PALETTE["gold"]], height=0.58)
    bar_ax.set_yticks(y, analyte_loss.index)
    bar_ax.invert_yaxis()
    bar_ax.set_xlabel("Support-attributable loss\n(sensor-hours)")
    bar_ax.set_xlim(0, max(hours) * 1.18)
    for index, value in enumerate(hours):
        bar_ax.text(value + max(hours) * 0.02, index, f"{int(value):,}", va="center", fontsize=6.0)
    style_axes(bar_ax)
    panel_label(bar_ax, "b")
    return _save_audit_figure(fig, figure_root, "FigD5_S3_L1_blocker_matrix")


def _figure_counterfactual(counterfactual: pd.DataFrame, figure_root: Path) -> list[Path]:
    configure_style()
    frame = counterfactual.copy()
    order = [
        "Current",
        "Family days repaired",
        "Family months repaired",
        "Node days repaired",
        "Node months repaired",
        "Node coverage repaired",
        "All L2 support repaired",
    ]
    frame["scenario"] = pd.Categorical(frame["scenario"], categories=order, ordered=True)
    frame = frame.sort_values("scenario")
    y = np.arange(len(frame))
    values = frame["report_coverage"].to_numpy() * 100.0
    colors = [
        PALETTE["gray"] if label not in {"Family days repaired", "All L2 support repaired"}
        else PALETTE["teal"]
        for label in frame["scenario"].astype(str)
    ]
    fig, ax = plt.subplots(figsize=(FIGURE_WIDTH_MM / 25.4, 3.75), constrained_layout=True)
    ax.hlines(y, 0, values, color=colors, linewidth=2.4)
    ax.scatter(values, y, color=colors, s=28, zorder=3)
    ax.set_yticks(y, frame["scenario"].astype(str))
    ax.invert_yaxis()
    ax.set_xlim(0, 100)
    ax.set_xlabel("Diagnostic report-eligible coverage (%)")
    ax.set_ylabel("Prespecified counterfactual")
    for index, value in enumerate(values):
        ax.text(min(value + 1.2, 98), index, f"{value:.1f}%", va="center",
                ha="left" if value < 96 else "right", fontweight="bold" if value > 0 else "normal")
    ax.axvline(values[-1], color=PALETTE["teal"], linestyle=":", linewidth=0.8)
    ax.text(0.91, 0.98, "Evidence-ready ceiling", transform=ax.transAxes,
            ha="right", va="top", color=PALETTE["teal"])
    style_axes(ax)
    return _save_audit_figure(fig, figure_root, "FigD5_S4_counterfactual_coverage_recovery")


def _write_report(path: Path, tables: dict[str, pd.DataFrame], metadata: dict[str, object]) -> None:
    shift = tables["01b_pre_post_regime_shift"]
    top = shift.sort_values("post_occupancy", ascending=False).iloc[0]
    blockers = tables["03_L1_to_L2_blockers"]
    counterfactual = tables["06_counterfactual_coverage"].set_index("scenario")
    attribution = tables["05_coverage_loss_attribution"].set_index("loss_class")
    maturity = tables["04_L2_to_L3_blockers"]
    horizon = tables["07_reference_horizon_sensitivity"]
    current = 100.0 * float(counterfactual.loc["Current", "report_coverage"])
    family = 100.0 * float(counterfactual.loc["Family days repaired", "report_coverage"])
    support_loss = int(attribution.loc["limited_support", "loss_sensor_hours"])
    support_pp = float(
        attribution.loc["limited_support", "coverage_percentage_point_contribution"]
    )
    ood_loss = int(attribution.loc["out_of_template", "loss_sensor_hours"])
    ood_pp = float(
        attribution.loc["out_of_template", "coverage_percentage_point_contribution"]
    )
    incomplete_loss = int(attribution.loc["not_evaluable", "loss_sensor_hours"])
    incomplete_pp = float(
        attribution.loc["not_evaluable", "coverage_percentage_point_contribution"]
    )
    family_far_blocked = int(maturity["family_far"].sum())
    node_far_blocked = int(maturity["node_far"].sum())
    horizon_r2 = horizon[
        horizon["regime_label"].eq(str(top["regime_label"]))
        & horizon["reference_fraction"].eq(0.80)
    ].iloc[0]
    report = f"""# D5 support-migration audit

## Scope and frozen boundary

- Audit ID: `{metadata['audit_id']}`
- Source run: `{metadata['source_run_id']}`
- Reference endpoint: `{metadata['reference_end']}`
- Post-reference analysis starts after the 7 d embargo: `{metadata['post_start']}`
- Authoritative D5 scores modified: **No**

## Confirmatory result

The decline in D5 report eligibility after January 2026 is **reference-horizon dominated**, not evidence of deteriorating sensor quality. {top['regime_label']} occupies {100.0 * float(top['post_occupancy']):.1f}% of the post-embargo period. All {len(blockers)} sensor-by-{top['regime_label']} templates are L1 because family effective support is 29 calendar days, below the prespecified L2 threshold of 40 days.

The competing explanations were not supported:

- distinct-month support is 2 months and passes the L2 contract;
- node effective support is 29 days and passes the 20-day threshold;
- node reconstruction coverage is 99.95%, far above the 0.60 threshold;
- stability, blocked holdouts and FAR do not participate in L1-to-L2 admission and therefore cannot explain the L1 migration.

Current post-embargo report coverage is {current:.2f}%. The mutually exclusive loss decomposition is {support_loss:,} limited-support sensor-hours ({support_pp:.2f} percentage points), {ood_loss:,} OOD/out-of-template sensor-hours ({ood_pp:.2f} points), and {incomplete_loss:,} not-evaluable sensor-hours ({incomplete_pp:.2f} points). A diagnostic counterfactual that repairs only family effective-day support increases coverage to {family:.2f}% (+{family-current:.2f} points), identical to the all-L2-support ceiling. It correctly preserves the residual {100.0-family:.2f}% OOD/incomplete-evidence loss.

## Scientific interpretation

The frozen K=4 context model partitions the study trajectory into temporally ordered operating regimes. R2 first becomes dominant shortly before the reference cutoff and reaches 100% occupancy in February-April. Consequently, R2 has only 29 complete reference calendar days even though its node reconstruction evidence is almost complete. D5 correctly abstains rather than extrapolating an under-mature family template.

This is an evidence-availability shift, not a low D5 score. Manuscript language should therefore state that the late-period estimand is limited by frozen-template maturity. Availability-aware and complete-evidence composites must remain separated.

The 0.80 reference-fraction shadow places {top['regime_label']} in {int(horizon_r2['occupied_calendar_days_upper_bound'])} occupied calendar days across {int(horizon_r2['distinct_months'])} months, which clears the *occupancy-horizon* L2 threshold. This is a descriptive upper bound, not a rebuilt effective-support result: high-quality family/node evidence, templates, validation and future performance were not recalculated. It supports a prospective full shadow refit but does not justify retroactively replacing the frozen 0.70 model.

Among the 42 templates already at L2/L3, family FAR blocks {family_far_blocked} and node FAR blocks {node_far_blocked} from L3. These constraints explain action-grade maturity, not the post-reference L1 migration.

## Recommended D5 actions

1. Keep the authoritative v2.4 scores and L2 thresholds unchanged.
2. Publish the support-migration audit as Supplementary/Extended Data evidence and carry `support_level`, `family_n_effective`, `reference_end`, and D5 availability into DQR aggregation metadata.
3. For a future prospective release, establish a rolling but versioned template lifecycle: a candidate R2 template may be promoted only after 40 effective days and at least 2 months, followed by a frozen prospective validation period.
4. Run a predeclared 0.80 reference-fraction shadow refit only as a future-version study; rebuild effective support, templates, blocked validation and OOD rather than treating occupied days as effective days.
5. Do not merge regimes solely to recover coverage. A K=3/K=5 refit is a new model-selection exercise and requires outer-fold discrimination, localization, OOD, and process interpretability checks.
6. Keep L2-to-L3 stability/FAR limitations separate; they constrain action-grade deployment but do not cause the observed report-coverage loss.

## Publication boundary

The audit is post hoc but uses prespecified support thresholds and frozen artifacts. Counterfactuals are diagnostic upper bounds, not alternative production scores. No post-reference observations were used to update the authoritative templates.
"""
    path.write_text(report, encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="Audit D5 post-reference L1 support migration")
    parser.add_argument("--d5-root", type=Path, default=D5_ROOT)
    args = parser.parse_args()
    output_root = args.d5_root / "outputs" / "audit" / "support_migration"
    figure_root = args.d5_root / "outputs" / "figures" / "supplementary" / "support_migration"
    output_root.mkdir(parents=True, exist_ok=True)
    figure_root.mkdir(parents=True, exist_ok=True)

    tables, metadata = D5SupportMigrationAudit(args.d5_root).run()
    artifacts = _write_table_files(output_root, tables)
    workbook = output_root / "D5_support_migration_audit.xlsx"
    _write_workbook(workbook, tables)
    artifacts.append(workbook)

    figure_artifacts: list[Path] = []
    figure_artifacts.extend(
        _figure_occupancy(
            tables["01_monthly_regime_occupancy"],
            figure_root,
            pd.Timestamp(metadata["reference_end"]),
        )
    )
    figure_artifacts.extend(_figure_maturity(tables["02_template_occupancy_56"], figure_root))
    figure_artifacts.extend(_figure_blockers(tables["03_L1_to_L2_blockers"], figure_root))
    figure_artifacts.extend(_figure_counterfactual(tables["06_counterfactual_coverage"], figure_root))
    artifacts.extend(figure_artifacts)

    report = output_root / "D5_SUPPORT_MIGRATION_AUDIT_REPORT.md"
    _write_report(report, tables, metadata)
    artifacts.append(report)
    manifest = output_root / "D5_support_migration_manifest.json"
    write_manifest(manifest, metadata, artifacts, base_root=args.d5_root)
    print(json.dumps({
        "audit_id": metadata["audit_id"],
        "reference_end": metadata["reference_end"],
        "post_start": metadata["post_start"],
        "output_root": str(output_root),
        "figure_root": str(figure_root),
        "artifacts": len(artifacts) + 1,
    }, indent=2))


if __name__ == "__main__":
    main()
