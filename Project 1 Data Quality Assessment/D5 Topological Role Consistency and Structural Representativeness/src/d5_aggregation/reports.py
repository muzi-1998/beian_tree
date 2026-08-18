from __future__ import annotations

from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def write_workbook(path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, frame in sheets.items():
            frame.to_excel(writer, sheet_name=name[:31], index=False)
        for sheet in writer.book.worksheets:
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            for cell in sheet[1]:
                cell.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="3D5A80")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for column_index, cells in enumerate(sheet.columns, 1):
                values = [str(cell.value) if cell.value is not None else "" for cell in cells]
                width = min(max(max(map(len, values), default=0) + 2, 10), 42)
                sheet.column_dimensions[get_column_letter(column_index)].width = width
                for cell in cells[1:]:
                    cell.font = Font(name="Arial", size=8)
                    cell.alignment = Alignment(vertical="top", wrap_text=True)


def _pct(value: float) -> str:
    return f"{100 * value:.1f}%"


def write_scientific_report(
    path: Path,
    *,
    run_id: str,
    node: pd.DataFrame,
    pair: pd.DataFrame,
    invariance: pd.DataFrame,
    aggregators: pd.DataFrame,
    weight_summary: pd.DataFrame,
    coverage_shift: pd.DataFrame,
    block_summary: pd.DataFrame,
    construct: pd.DataFrame,
    pending: pd.DataFrame,
) -> None:
    node_counts = node["coverage_class"].value_counts()
    pair_counts = pair["coverage_class"].value_counts()
    primary = block_summary.loc[block_summary["block_hours"].eq(168)].copy()
    selection = coverage_shift.loc[
        (coverage_shift["stratum_type"] == "overall")
        & (coverage_shift["stratum"] == "all")
    ].iloc[0]
    d4_d5 = construct.loc[
        (construct["scope"] == "D4_D5_dual_scope")
        & (construct["right"] == "D5_report")
    ].iloc[0]
    node_equal = aggregators.loc[
        (aggregators["scope"] == "node")
        & (aggregators["aggregator"] == "arithmetic_equal_weight")
    ].iloc[0]
    pair_equal = aggregators.loc[
        (aggregators["scope"] == "pair")
        & (aggregators["aggregator"] == "arithmetic_equal_weight")
    ].iloc[0]
    node_weight = weight_summary.loc[
        (weight_summary["scope"] == "node")
        & (weight_summary["metric"] == "spearman_vs_equal")
    ].iloc[0]
    pair_weight = weight_summary.loc[
        (weight_summary["scope"] == "pair")
        & (weight_summary["metric"] == "spearman_vs_equal")
    ].iloc[0]
    rows = [
        "# D1-D5 hierarchical data-quality aggregation report",
        "",
        f"Run ID: `{run_id}`",
        "",
        "## Confirmatory estimands",
        "",
        "The release implements a hierarchical Quality-Evidence-Gate contract. "
        "D1, D2 and eligible D5 evidence form node-level quality; homologous left/right "
        "node quality and native D4_raw form pair-level quality. D3 is retained as an "
        "independent non-compensatory safety gate and is never averaged into either score. "
        "Evidence completeness is reported separately and is never multiplied by quality.",
        "",
        "- `Q_node_full = mean(D1, D2, D5_report)` when all three formal scores exist.",
        "- `Q_node_available = mean(D1, D2[, D5_report])`; D1 and D2 are mandatory.",
        "- `Q_pair_full = mean(left Q_node_full, right Q_node_full, D4_raw)`.",
        "- `Q_pair_available = mean(left Q_node_available, right Q_node_available, D4_raw)`.",
        "- `E_node` and `E_pair` quantify evidence coverage, not data quality.",
        "",
        "## Current results",
        "",
        f"The node table contains {len(node):,} sensor-hours: "
        f"{node_counts.get('full', 0):,} Full, {node_counts.get('basic', 0):,} Basic, "
        f"{node_counts.get('limited', 0):,} Limited and "
        f"{node_counts.get('insufficient', 0):,} Insufficient. The pair table contains "
        f"{len(pair):,} native pair-hours: {pair_counts.get('full', 0):,} Full, "
        f"{pair_counts.get('basic', 0):,} Basic and {pair_counts.get('limited', 0):,} Limited.",
        "",
        f"The complete-case identity check passed for both levels (maximum absolute "
        f"difference {invariance['maximum_absolute_difference'].max():.3g}). Under equal "
        f"arithmetic aggregation, the complete-evidence low-tail rate was "
        f"{_pct(node_equal['low_tail_rate'])} for nodes and {_pct(pair_equal['low_tail_rate'])} "
        "for pairs.",
        "",
        f"The Full and Basic node subsets are not exchangeable: their overall standardized "
        f"mean difference was {selection['standardized_mean_difference_full_minus_basic']:.2f} "
        f"and Wasserstein distance was {selection['wasserstein_distance']:.2f}. Full is "
        "therefore the complete-evidence estimand; availability-aware Basic extends coverage "
        "but must be displayed separately.",
        "",
        f"Across constrained prespecified weights, median Spearman agreement with equal "
        f"weights was {node_weight['median']:.3f} for nodes and {pair_weight['median']:.3f} "
        "for pairs. This supports rank robustness within the examined weight region, but does "
        "not identify optimal weights because no frozen downstream criterion is available.",
        "",
        f"D4 and formal D5_report showed modest association (Spearman rho = "
        f"{d4_d5['spearman']:.3f}, 7 d synchronized-block 95% CI "
        f"{d4_d5['spearman_ci_low']:.3f} to {d4_d5['spearman_ci_high']:.3f}); low-tail "
        f"Jaccard was {d4_d5['low_tail_jaccard']:.3f}. This supports complementarity, not "
        "causal independence.",
        "",
        "## Statistical interpretation",
        "",
        "The primary uncertainty analysis uses synchronized 7 d process-time blocks so that "
        "all sensors and pairs sharing an operating disturbance remain in the same resample. "
        "The 48 h and 14 d analyses are sensitivity bounds. Inferential intervals are only "
        "reported when at least six independent blocks are available.",
        "",
    ]
    for _, item in primary.iterrows():
        rows.append(
            f"- {item['scope']} / {item['estimand']}: {item['estimate']:.3f} "
            f"(95% CI {item['ci_low']:.3f}-{item['ci_high']:.3f}; "
            f"{int(item['n_evaluable_plant_hours']):,} evaluable plant-hours)."
        )
    rows.extend(
        [
            "",
            "## Claim boundary and release decision",
            "",
            "This release is suitable for retrospective scientific aggregation and manuscript "
            "analysis. It is not an automated deployment release. D5 hard Veto remains disabled "
            "because controlled perturbation Top-1 localization is 0.767, below the prespecified "
            "0.80 criterion. A-E grades remain disabled. The D1 export lacks a distinct validated "
            "hard-fault interface, so Strict eligibility remains a contract candidate rather than "
            "a finalized automatic release flag.",
            "",
            "The prospective 2026-04-14 to 2026-07-31 holdout and downstream fitness-for-use "
            "validation remain pending because the required frozen D1-D5 and endpoint bundles do "
            "not exist. Missing maintenance/metrological evidence is recorded as not available and "
            "is never assigned a neutral or low score.",
            "",
            "## Pending registry",
            "",
        ]
    )
    for _, item in pending.iterrows():
        rows.append(
            f"- `{item['validation_id']}`: **{item['status']}**. "
            f"{item['reason']} Consequence: {item['blocking_effect']}."
        )
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(rows) + "\n", encoding="utf-8")


def write_directory_guide(path: Path) -> None:
    text = """# D1-D5 aggregation directory guide

The aggregation module is intentionally housed under D5 because D5 is the final
sensor-level evidence dimension, while the output remains a distinct D1-D5
integration layer. No native D1-D5 scores are overwritten.

## Configuration and code

- `configs/aggregation_v2.yaml`: frozen input hashes, estimands and validation rules.
- `src/d5_aggregation/`: loading, aggregation, statistics, figures, reports and manifests.
- `scripts/run_dqr_aggregation.py`: complete deterministic release build.
- `scripts/verify_dqr_aggregation.py`: formula, freshness, manifest and figure verification.
- `tests/test_aggregation_v2.py`: synthetic and released-output contract tests.

## Generated outputs

- `outputs/aggregation_v2/data/`: dimension-long, node-hour, pair-hour and monthly coverage tables.
- `outputs/aggregation_v2/validation/`: statistical workbooks and machine-readable QA.
- `outputs/aggregation_v2/figures/`: 183 mm Nature-style PNG/PDF/SVG/TIFF files.
- `outputs/aggregation_v2/source_data/`: one source-data workbook per figure.
- `outputs/aggregation_v2/reports/`: scientific report, captions and this guide.
- `outputs/aggregation_v2/manifests/`: frozen run and publication manifests.

Figures 6 and 7 specified in the study plan are intentionally absent until the
prospective holdout and downstream endpoint bundles become available. Their
absence is a prespecified pending status, not a missing build artifact.
"""
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")


def write_figure_captions(path: Path) -> None:
    text = """# Figure captions

**Figure 1 | Hierarchical quality, evidence completeness and safety-gate contract.**
Node quality combines D1, D2 and formally eligible D5 evidence, while pair quality
combines the two homologous nodes with native D4 evidence. D3 is a separate
non-compensatory gate. Quality and evidence completeness occupy separate axes.

**Figure 2 | Evidence availability and estimand stability across the study period.**
Monthly Full, Basic and Limited coverage is shown together with D5 availability,
plant-level Full and availability-aware scores, and evidence completeness. Full
and availability-aware series are distinct estimands and must not be pooled.

**Figure 3 | Pairwise construct complementarity across D1-D5.**
Pairwise-complete Spearman association and low-tail Jaccard overlap are estimated
without treating absence as a low score. D4-D5 intervals use synchronized 7 d
process-time block resampling; stratified estimates are descriptive when fewer
than six independent blocks are available.

**Figure 4 | Prespecified aggregation robustness.**
Equal arithmetic averaging is the primary method. Geometric, soft-minimum and
hard-minimum operators, constrained weight draws and leave-one-dimension-out
analyses are sensitivity checks and do not replace the primary score.

**Figure 5 | Multidimensional evidence around representative sensor and pair episodes.**
Hourly D1, D2, D5, node and native D4 evidence are shown around prespecified or
algorithmically selected cases. Grey shading denotes D3 Warn and hatching denotes
NotEvaluated. The figure does not claim maintenance-confirmed fault truth.

Figures 6 and 7 remain pending because prospective holdout scores and frozen
downstream endpoint bundles are unavailable.
"""
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")
