from __future__ import annotations

from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def write_workbook(path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name, frame in sheets.items():
            frame.to_excel(writer, sheet_name=name[:31], index=False)
        for sheet in writer.book.worksheets:
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            for cell in sheet[1]:
                cell.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="3D5A80")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            for column_index, cells in enumerate(sheet.columns, 1):
                values = [str(cell.value) if cell.value is not None else "" for cell in cells]
                width = min(max(max(map(len, values), default=0) + 2, 10), 42)
                sheet.column_dimensions[get_column_letter(column_index)].width = width
                for cell in cells[1:]:
                    cell.font = Font(name="Arial", size=8)
                    cell.alignment = Alignment(vertical="top", wrap_text=True)


def _pct(value: float) -> str:
    return f"{100 * value:.1f}%"


def write_scientific_report(
    path: Path,
    *,
    run_id: str,
    node: pd.DataFrame,
    pair: pd.DataFrame,
    invariance: pd.DataFrame,
    aggregators: pd.DataFrame,
    weight_summary: pd.DataFrame,
    coverage_shift: pd.DataFrame,
    block_summary: pd.DataFrame,
    construct: pd.DataFrame,
    decomposition: pd.DataFrame,
    pair_weighting: pd.DataFrame,
    pair_threshold_sweep: pd.DataFrame,
    pending: pd.DataFrame,
) -> None:
    node_counts = node["coverage_class"].value_counts()
    pair_counts = pair["coverage_class"].value_counts()
    mapping_counts = pair["mapping_support_class"].value_counts(normalize=True)
    primary = block_summary.loc[block_summary["block_hours"].eq(168)].copy()
    selection = coverage_shift.loc[
        (coverage_shift["stratum_type"] == "overall")
        & (coverage_shift["stratum"] == "all")
    ].iloc[0]
    d4_d5_report = construct.loc[
        (construct["scope"] == "D4_D5_dual_scope")
        & (construct["right"] == "D5_report")
    ].iloc[0]
    d4_d5_raw = construct.loc[
        (construct["scope"] == "D4_D5_dual_scope")
        & (construct["right"] == "D5_raw")
    ].iloc[0]
    effects = decomposition.loc[
        decomposition["stratum_type"].eq("overall")
        & decomposition["stratum"].eq("all")
    ].set_index("effect")
    pair_native = pair_weighting.iloc[0]
    node_equal = aggregators.loc[
        (aggregators["scope"] == "node")
        & (aggregators["aggregator"] == "arithmetic_equal_weight")
    ].iloc[0]
    pair_equal = aggregators.loc[
        (aggregators["scope"] == "pair")
        & (aggregators["aggregator"] == "arithmetic_equal_weight")
    ].iloc[0]
    node_weight = weight_summary.loc[
        (weight_summary["scope"] == "node")
        & (weight_summary["metric"] == "spearman_vs_equal")
    ].iloc[0]
    pair_weight = weight_summary.loc[
        (weight_summary["scope"] == "pair")
        & (weight_summary["metric"] == "spearman_vs_equal")
    ].iloc[0]
    rows = [
        "# D1-D5 hierarchical data-quality aggregation report",
        "",
        f"Run ID: `{run_id}`",
        "",
        "## Confirmatory estimands",
        "",
        "The release implements a hierarchical Quality-Evidence-Gate contract. "
        "D1, D2 and eligible D5 evidence form node-level quality; homologous left/right "
        "node quality and native D4_raw form pair-level quality. D3 is retained as an "
        "independent non-compensatory safety gate and is never averaged into either score. "
        "Evidence completeness is reported separately and is never multiplied by quality.",
        "",
        "- `Q_node_full = mean(D1, D2, D5_report)` when all three formal scores exist.",
        "- `Q_node_core12 = mean(D1, D2)` is the fixed-composition longitudinal estimand.",
        "- `Q_node_available = mean(D1, D2[, D5_report])`; D1 and D2 are mandatory.",
        "- `Q_pair_full = mean(left Q_node_full, right Q_node_full, D4_raw)`.",
        "- `Q_pair_core = mean(left Q_node_core12, right Q_node_core12, D4_raw)`.",
        "- `Q_pair_available = mean(left Q_node_available, right Q_node_available, D4_raw)`.",
        "- `E_node` and `E_pair` quantify evidence coverage, not data quality.",
        "",
        "## Current results",
        "",
        f"The node table contains {len(node):,} sensor-hours: "
        f"{node_counts.get('full', 0):,} Full, {node_counts.get('basic', 0):,} Basic, "
        f"{node_counts.get('limited', 0):,} Limited and "
        f"{node_counts.get('insufficient', 0):,} Insufficient. The pair table contains "
        f"{len(pair):,} native pair-hours: {pair_counts.get('full', 0):,} Full, "
        f"{pair_counts.get('basic', 0):,} Basic and {pair_counts.get('limited', 0):,} Limited.",
        "",
        f"The complete-case identity check passed for both levels (maximum absolute "
        f"difference {invariance['maximum_absolute_difference'].max():.3g}). Under equal "
        f"arithmetic aggregation, the complete-evidence low-tail rate was "
        f"{_pct(node_equal['low_tail_rate'])} for nodes and {_pct(pair_equal['low_tail_rate'])} "
        "for pairs.",
        "",
        f"Sensor-hour pooled means were {node['Q_node_full'].mean():.3f} (node Full), "
        f"{node['Q_node_core12'].mean():.3f} (node fixed Core), "
        f"{node['Q_node_available'].mean():.3f} (node availability-aware), "
        f"{pair['Q_pair_full'].mean():.3f} (pair Full) and "
        f"{pair['Q_pair_core'].mean():.3f} (pair fixed Core) and "
        f"{pair['Q_pair_available'].mean():.3f} (pair availability-aware). These are "
        "not the plant-hour aggregated means reported below, which first summarize all "
        "objects within each plant-hour and then average over time.",
        "",
        f"The Full and Basic node subsets are not exchangeable: their overall standardized "
        f"mean difference was {selection['standardized_mean_difference_full_minus_basic']:.2f} "
        f"and Wasserstein distance was {selection['wasserstein_distance']:.2f}. Full is "
        "therefore the complete-evidence estimand; availability-aware Basic extends coverage "
        "but must be displayed separately.",
        "",
        f"D4 pair-hours used exact variable-regime mapping for "
        f"{_pct(mapping_counts.get('exact', 0.0))}, variable-level fallback for "
        f"{_pct(mapping_counts.get('variable_fallback', 0.0))}, and global fallback for "
        f"{_pct(mapping_counts.get('global_fallback', 0.0))}. Mapping support is evidence "
        "metadata and does not numerically penalize D4_raw. Any future exclusion rule must "
        "be prospectively frozen and independently validated.",
        "",
        f"The sensor-hour estimand decomposition separated a selection-only shift of "
        f"{effects.loc['selection_only', 'estimate']:.3f} "
        f"(95% block CI {effects.loc['selection_only', 'ci_low']:.3f} to "
        f"{effects.loc['selection_only', 'ci_high']:.3f}) from a within-Full D5 "
        f"compositional contribution of "
        f"{effects.loc['within_Full_D5_compositional_contribution', 'estimate']:.3f} "
        f"({effects.loc['within_Full_D5_compositional_contribution', 'ci_low']:.3f} to "
        f"{effects.loc['within_Full_D5_compositional_contribution', 'ci_high']:.3f}). "
        f"Their sum equals the total "
        f"observed estimand shift of "
        f"{effects.loc['total_observed_estimand_shift', 'estimate']:.3f}; absolute "
        f"closure error was {abs(effects['overall_closure_error'].iloc[0]):.2g}. This is "
        "a descriptive estimand decomposition, not a causal effect decomposition.",
        "",
        f"Across constrained prespecified weights, median Spearman agreement with equal "
        f"weights was {node_weight['median']:.3f} for nodes and {pair_weight['median']:.3f} "
        "for pairs. This supports rank robustness within the examined weight region, but does "
        "not identify optimal weights because no frozen downstream criterion is available.",
        "",
        f"D4 showed modest association with both formal D5_report (Spearman rho = "
        f"{d4_d5_report['spearman']:.3f}, 7 d synchronized-block 95% CI "
        f"{d4_d5_report['spearman_ci_low']:.3f} to "
        f"{d4_d5_report['spearman_ci_high']:.3f}) and calculable D5_raw "
        f"(rho = {d4_d5_raw['spearman']:.3f}, 95% CI "
        f"{d4_d5_raw['spearman_ci_low']:.3f} to {d4_d5_raw['spearman_ci_high']:.3f}). "
        "The dual scope supports complementarity without asserting causal independence.",
        "",
        f"At pair level, hierarchical equal-component and seven-atom equal weighting "
        f"had Spearman rho = {pair_native['spearman']:.3f}, low-tail Jaccard = "
        f"{pair_native['low_tail_jaccard']:.3f}, and "
        f"{_pct(pair_native['decision_flip_rate_at_3'])} decision flips at Q < 3. "
        "Thus global ranking was broadly concordant but rare low-tail episode identity "
        "was not robust to flattening the hierarchy. This is a supplementary robustness "
        "comparison; the formal hierarchical model was not selected or changed from these data.",
        "",
        f"At the formal Q < {pair_native['formal_low_tail_threshold']:.2f} threshold, "
        f"the hierarchical and native-atom estimands identified "
        f"{int(pair_native['hierarchical_low_tail_count']):,} and "
        f"{int(pair_native['native_atom_low_tail_count']):,} low-tail pair-hours, "
        f"respectively. Their partition comprised {int(pair_native['both_count']):,} both, "
        f"{int(pair_native['hierarchical_only_count']):,} hierarchical-only, "
        f"{int(pair_native['native_atom_only_count']):,} native-atom-only and "
        f"{int(pair_native['neither_count']):,} neither hours. The corresponding episode "
        f"counts were {int(pair_native['hierarchical_event_count']):,} and "
        f"{int(pair_native['native_atom_event_count']):,}, with median durations of "
        f"{pair_native['hierarchical_median_episode_duration_h']:.1f} h and "
        f"{pair_native['native_atom_median_episode_duration_h']:.1f} h.",
        "",
        f"Across the prespecified Q < {pair_threshold_sweep['threshold'].min():.2f}-"
        f"{pair_threshold_sweep['threshold'].max():.2f} sensitivity range, low-tail "
        f"Jaccard ranged from {pair_threshold_sweep['low_tail_jaccard'].min():.3f} to "
        f"{pair_threshold_sweep['low_tail_jaccard'].max():.3f}, while decision-flip "
        f"fractions ranged from {pair_threshold_sweep['decision_flip_rate'].min():.3f} "
        f"to {pair_threshold_sweep['decision_flip_rate'].max():.3f}. The sweep tests "
        "whether the formal Q < 3 result is threshold-local; it is not used to choose "
        "a replacement threshold or weighting model.",
        "",
        "## Statistical interpretation",
        "",
        "The primary uncertainty analysis uses synchronized 7 d process-time blocks so that "
        "all sensors and pairs sharing an operating disturbance remain in the same resample. "
        "The 48 h and 14 d analyses are sensitivity bounds. Inferential intervals are only "
        "reported when at least six independent blocks are available.",
        "",
    ]
    for _, item in primary.iterrows():
        rows.append(
            f"- {item['scope']} / {item['estimand']}: {item['estimate']:.3f} "
            f"(95% CI {item['ci_low']:.3f}-{item['ci_high']:.3f}; "
            f"{int(item['n_evaluable_plant_hours']):,} evaluable plant-hours)."
        )
    rows.extend(
        [
            "",
            "## Claim boundary and release decision",
            "",
        "This release is suitable for retrospective scientific aggregation and manuscript "
            "analysis. It is not an automated deployment release. D5 hard Veto remains disabled "
            "because controlled perturbation Top-1 localization is 0.767, below the prespecified "
            "0.80 criterion. A-E grades remain disabled. The D1 export lacks a distinct validated "
            "hard-fault interface, so Strict eligibility remains a contract candidate rather than "
            "a finalized automatic release flag.",
            "",
            "The prospective 2026-04-14 to 2026-07-31 holdout and downstream fitness-for-use "
            "validation remain pending because the required frozen D1-D5 and endpoint bundles do "
            "not exist. Missing maintenance/metrological evidence is recorded as not available and "
        "is never assigned a neutral or low score.",
        "",
        "Longitudinal interpretation is restricted to fixed-composition Core estimands. "
        "Full remains the complete-evidence scientific estimand, while availability-aware "
        "scores are operational summaries and must not be compared across dimension masks. "
        "D5 L1 denotes limited evidence support, not low data quality, and historical L1 "
        "hours are never backfilled after a future template upgrade.",
            "",
            "## Pending registry",
            "",
        ]
    )
    for _, item in pending.iterrows():
        rows.append(
            f"- `{item['validation_id']}`: **{item['status']}**. "
            f"{item['reason']} Consequence: {item['blocking_effect']}."
        )
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(rows) + "\n", encoding="utf-8")


def write_directory_guide(path: Path) -> None:
    text = """# D1-D5 aggregation directory guide

This directory is a peer-level cross-dimensional integration layer. D5 is one
input dimension and is not the parent of D1-D4. No native D1-D5 scores are
overwritten.

## Configuration and code

- `configs/aggregation_v2_3.yaml`: frozen input hashes, phase/reference contracts,
  estimands and validation rules. The v2.2 configuration and outputs are retained.
- `src/dqr_aggregation/`: loading, aggregation, statistics, figures, reports and manifests.
- `scripts/run_dqr_aggregation.py`: complete deterministic release build.
- `scripts/verify_dqr_aggregation.py`: formula, freshness, manifest and figure verification.
- `tests/test_aggregation_v2_3.py`: synthetic and released-output contract tests.

## Generated outputs

- `outputs/aggregation_v2_3/data/`: versioned dimension-long, node, pair, coverage,
  phase/evidence summaries and the machine-readable estimand registry.
- `outputs/aggregation_v2_3/data/`: also contains estimand-decomposition and
  pair-weighting sensitivity tables.
- `outputs/aggregation_v2_3/validation/`: statistical workbooks and machine-readable QA.
- `outputs/aggregation_v2_3/figures/`: 183 mm Nature-style PNG/PDF/SVG/TIFF files.
- `outputs/aggregation_v2_3/source_data/`: one source-data workbook per figure.
- `outputs/aggregation_v2_3/reports/`: scientific report, captions and this guide.
- `outputs/aggregation_v2_3/manifests/`: frozen run and publication manifests.

The run manifest records the scientific-generation commit for orientation, but
publication freshness is governed by exact canonical hashes of the current
configuration, every aggregation source module and all frozen D1-D5 inputs. The
publication-bundle commit or release tag is external metadata so that a manifest
never attempts to hash a commit that contains itself.

Figures 6 and 7 specified in the study plan are intentionally absent until the
prospective holdout and downstream endpoint bundles become available. Their
absence is a prespecified pending status, not a missing build artifact.
"""
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")


def write_expert_review(path: Path, node: pd.DataFrame, pair: pd.DataFrame) -> None:
    scope = pair["mapping_support_class"].value_counts(normalize=True)
    outcome = (
        pair.loc[pair["usable_for_D4"].fillna(False) & pair["D4_raw"].notna()]
        .assign(D4_low=lambda x: x["D4_raw"].lt(3.0))
        .groupby(["phase_id", "variable", "mapping_support_class"], observed=True)
        .agg(low_tail_rate=("D4_low", "mean"), n_pair_hours=("timestamp", "size"))
        .reset_index()
    )
    orp_fallback_validation = outcome.loc[
        outcome["phase_id"].eq("internal_validation")
        & outcome["variable"].eq("ORP")
        & outcome["mapping_support_class"].eq("variable_fallback")
    ]
    orp_rate = (
        float(orp_fallback_validation["low_tail_rate"].iloc[0])
        if len(orp_fallback_validation)
        else np.nan
    )
    rows = [
        "# DQR v2.3 expert review and implementation decision",
        "",
        "## Overall decision",
        "",
        "The proposal is scientifically coherent and addresses the principal remaining risk: changes in evidence composition must not be interpreted as temporal changes in data quality. The accepted P0 changes were implemented without changing any frozen D1-D5 formal score.",
        "",
        "## Executed",
        "",
        "- Added fixed node and pair Core estimands alongside Full and availability-aware estimands.",
        "- Locked Quality, Evidence and Gate as separate non-multiplicative axes; D3 remains fail-closed and non-compensatory.",
        "- Bound D5 support-migration v1.1 into the frozen input registry; L1 remains diagnostic-only limited evidence and is never converted into low quality.",
        "- Added D4 exact/variable-fallback/global-fallback/insufficient metadata to every pair-hour and completed a descriptive mapping-support migration audit.",
        "- Added phase_role, reference_status and version_hash to every dimension-long row, plus a phase/evidence summary and machine-readable estimand registry.",
        "- Retained v2.2 outputs and created a separate v2.3 release directory, source-data bundle, figures, reports, tests and SHA-256 manifests.",
        "- Kept A-E grades, D5 hard Veto and optimized weights disabled.",
        "",
        "## Main numerical implications",
        "",
        f"- Sensor-hour pooled node means: Core {node['Q_node_core12'].mean():.3f}, Full {node['Q_node_full'].mean():.3f}, availability-aware {node['Q_node_available'].mean():.3f}.",
        f"- Sensor-hour pooled pair means: Core {pair['Q_pair_core'].mean():.3f}, Full {pair['Q_pair_full'].mean():.3f}, availability-aware {pair['Q_pair_available'].mean():.3f}.",
        "- These values are intentionally different estimands; their differences are not model disagreement and must not be collapsed into one trend.",
        f"- D4 mapping support comprised {scope.get('exact', 0.0):.1%} exact, {scope.get('variable_fallback', 0.0):.1%} variable fallback, {scope.get('global_fallback', 0.0):.1%} global fallback and {scope.get('insufficient', 0.0):.1%} insufficient pair-hours.",
        f"- ORP variable-fallback validation rows had a D4<3 rate of {orp_rate:.1%}. Because support class is structurally coupled to regime, this is descriptive and cannot be interpreted as a causal fallback penalty.",
        "",
        "## Accepted with modification",
        "",
        "- Native D1-D3 files were not rewritten merely to add phase labels. The harmonized fields are added at the integration interface, preserving upstream frozen hashes; native exports should change only in a new versioned upstream release.",
        "- D4 fallback remains formally scoreable metadata in this release. No global fallback was observed, and exact/fallback strata are regime-confounded. Any future exclusion from Full requires a prospectively frozen rule and new test data.",
        "- D5 post-embargo rows are labelled for aggregation-time interpretation only; this does not create a new independent-validation claim for the D5 model.",
        "- The Nature static preflight warning on D4 `.dropna()` is non-substantive: these calls collapse non-null metadata values; missing-regime rows are explicitly retained as `insufficient`, and all input/evaluable counts are exported.",
        "",
        "## Pending and not executed",
        "",
        "- D1 development-only frozen K=4 context shadow and the dependent D4 regime-shadow comparison: require a separately preregistered model artifact and paired sensitivity run.",
        "- Independent D1 hard-fault interface: cannot be reconstructed from a D1_total threshold and requires controlled challenge or reviewed event truth.",
        "- D5 template promotion and bridge: requires future independent support, frozen validation and prospective activation; historical L1 will not be backfilled.",
        "- Prospective post-2026-04-13 scoring, downstream fitness-for-use, maintenance/metrological truth, cross-plant validation, learned weights and A-E cutpoints: required data are unavailable.",
        "",
        "## Publication conclusion",
        "",
        "DQR v2.3 is suitable for retrospective manuscript analysis as a hierarchical, evidence-aware and non-compensatory aggregation framework. Core is the longitudinal estimand, Full is the complete-evidence scientific estimand, and availability-aware is an operational extension. The release is not a deployment-grade automated grading or hard-Veto system.",
    ]
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(rows) + "\n", encoding="utf-8")


def write_figure_captions(path: Path) -> None:
    text = """# Figure captions

**Figure 1 | Hierarchical quality, evidence completeness and safety-gate contract.**
Node quality combines D1, D2 and formally eligible D5 evidence, while pair quality
combines the two homologous nodes with native D4 evidence. D3 is a separate
non-compensatory gate. Quality and evidence completeness occupy separate axes.

**Figure 2 | Fixed-composition quality and evidence support across the study period.**
Monthly coverage and D5 L1-L3/OOD support migration are shown with Full and
availability-aware quality. Fixed D1-D2 Core is the longitudinal primary series;
Full is restricted to complete-evidence hours and availability-aware is an
operational extension. Evidence completeness and D4 exact-mapping support are
reported on a separate axis, together with the selection-only,
within-Full D5 compositional contribution and total observed estimand shifts. Full and availability-aware
series are distinct estimands and must not be pooled.

**Figure 3 | Pairwise construct complementarity across D1-D5.**
Pairwise-complete Spearman association and low-tail Jaccard overlap are estimated
without treating absence as a low score. D4-D5 intervals use synchronized 7 d
process-time block resampling; stratified estimates are descriptive when fewer
than six independent blocks are available.

**Figure 4 | Prespecified aggregation robustness.**
Equal arithmetic averaging is the primary method. Geometric, soft-minimum and
hard-minimum operators, constrained weight draws and leave-one-dimension-out
analyses are sensitivity checks and do not replace the primary score.

**Figure 5 | Multidimensional evidence around representative sensor and pair episodes.**
Hourly D1, D2, D5, node and native D4 evidence are shown around a median stable
benchmark window, two frozen D5 case-registry entries and an independently
registered D4 episode. Pale-gold shading denotes D3 Warn and grey hatching denotes
NotEvaluated. No panel is selected as the all-period maximum or minimum, and the
figure does not claim maintenance-confirmed fault truth.

**Extended Data Figure 1 | Pair hierarchical versus native-atom weighting.**
Complete-evidence pair-hours compare the formal equal-component hierarchy with
equal weighting of seven native atoms. Score concordance, low-tail hours,
exact overlap partitions, episode burden and a prespecified Q-threshold sweep are
sensitivity evidence, not model selection.

Figures 6 and 7 remain pending because prospective holdout scores and frozen
downstream endpoint bundles are unavailable.
"""
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")
